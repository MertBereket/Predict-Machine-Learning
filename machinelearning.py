import pandas as pd
from openpyxl import load_workbook
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score
from sklearn.model_selection import train_test_split

arac = load_workbook("ham_veri_dolu.xlsx")
sheet = arac.active
sheet = arac['ham veri']

dataset = pd.read_excel('ham_veri_dolu.xlsx')

X = dataset[['Yıl', 'Motor Hacmi', 'Motor Gücü', 'Kilometre']]
y = dataset[['Fiyat']]
arac = load_workbook("ham_veri_dolu.xlsx")

###Train Test

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=0)

###LinearRegresyon
regressor = LinearRegression()
regressor.fit(X_train, y_train)

y_pred = regressor.predict(X_test)


def pred(yil, motor, hp, km):
    return regressor.predict(
        [
            [yil, motor, hp, km]
        ]
    )[0][0]


###
# print(X[0])
test = pred(2020, 1968, 190, 0)
print(test)
RKare = r2_score(y_test, y_pred)
print(RKare)
