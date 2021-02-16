import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from openpyxl import Workbook, load_workbook

arac = load_workbook("ham veri.xlsx")
sheet = arac.active
sheet = arac['işlenmiş veri']

dataset = pd.read_excel('ham veri.xlsx')
X = dataset.iloc[:, :-1].values
y = dataset.iloc[:, -1].values

from sklearn.impute import SimpleImputer

imputer = SimpleImputer(missing_values=np.nan, strategy='mean')
imputer.fit(X[:, 5:7])
X[:, 5:7] = imputer.transform(X[:, 5:7])
print(X)

from sklearn.compose import ColumnTransformer
from sklearn.preprocessing import OneHotEncoder

ct = ColumnTransformer(transformers=[('encoder', OneHotEncoder(), [0, 1, 3, 4])], remainder='passthrough')
X = np.array(ct.fit_transform(X))
print(X)

from sklearn.model_selection import train_test_split

X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=1)
print(X_train)
print(X_test)
print(y_train)
print(y_test)

from sklearn.preprocessing import StandardScaler

sc = StandardScaler()
X_train[:, 8:12] = sc.fit_transform(X_train[:, 8:12])
X_test[:, 8:12] = sc.fit_transform(X_test[:, 8:12])
print(X_train)
print(X_test)

for i in range(1, len(X_train) + 1):
    for j in range(1, 13):
        sheet.cell(row=i, column=j).value = X_train[i - 1, j - 1]

for i in range(1, len(X_test) + 1):
    sheet.append(
        [X_test[i - 1, 0], X_test[i - 1, 1], X_test[i - 1, 2], X_test[i - 1, 3], X_test[i - 1, 4], X_test[i - 1, 5],
         X_test[i - 1, 6], X_test[i - 1, 7], X_test[i - 1, 8], X_test[i - 1, 9], X_test[i - 1, 10], X_test[i - 1, 11]])

for i in range(1, len(y_train)):
    sheet.cell(row=i, column=12).value = y_train[i]

arac.save("ham veri.xlsx")
arac.close()
